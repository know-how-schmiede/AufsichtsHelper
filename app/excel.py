import hashlib
import re
import zipfile
from datetime import date, datetime, time, timedelta
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

EXPECTED_COLUMNS = [
    "Pr\u00fcfungsname",
    "Datum",
    "Startzeit",
    "Dauer",
    "Pr\u00fcfer",
    "Aufsicht",
    "Abl\u00f6sung",
    "Raum",
]

COLUMN_ALIASES = {
    "Pr\u00fcfungsname": [
        "Pr\u00fcfungsname",
        "Fach",
        "Modul LV-Nr.",
        "Modul",
        "LV-Nr.",
        "EDV-Nr.",
    ],
    "Datum": ["Datum", "Pr\u00fcfungstag", "Tag"],
    "Startzeit": ["Startzeit", "Uhrzeit"],
    "Dauer": ["Dauer"],
    "Pr\u00fcfer": ["Pr\u00fcfer"],
    "Aufsicht": ["Aufsicht"],
    "Abl\u00f6sung": [
        "Abl\u00f6sung",
        "Abl\u00f6sung/ Beisitzer",
        "Abl\u00f6sung/Beisitzer",
    ],
    "Raum": ["Raum", "R\u00e4ume", "R\u00e4ume vorgezogen"],
}


def read_excel(file_path):
    rows = None
    try:
        rows = _read_rows_openpyxl(file_path)
    except Exception as exc:
        try:
            rows = _read_rows_fallback(file_path)
        except Exception:
            if isinstance(exc, InvalidFileException):
                raise ValueError(
                    "Excel-Datei konnte nicht gelesen werden (defektes XML). "
                    "Bitte die Datei in Excel/LibreOffice oeffnen und als .xlsx neu speichern."
                ) from exc
            raise ValueError(f"Excel-Datei konnte nicht gelesen werden: {exc}") from exc

    if not rows:
        raise ValueError("Excel-Datei ist leer")

    header_row_index, headers, score = _detect_header_row(rows)
    if score <= 0:
        raise ValueError(
            "Keine gueltige Kopfzeile gefunden. Bitte pruefen, ob die erste Zeile "
            "die Spaltenueberschriften enthaelt."
        )

    missing = [
        field for field in EXPECTED_COLUMNS if not _resolve_indices(headers, field)
    ]
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    data_rows = []
    for row in rows[header_row_index + 1 :]:
        record = _build_record(headers, row)
        if record is None:
            continue
        data_rows.append(record)

    return data_rows


def normalize_name(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split()).casefold()


def split_names(value):
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []

    if any(sep in text for sep in [";", "/", "\n", "\r"]):
        parts = re.split(r"[;/\n\r]+", text)
        return [part.strip() for part in parts if part and part.strip()]

    if text.count(",") >= 3:
        parts = [part.strip() for part in text.split(",") if part.strip()]
        if len(parts) % 2 == 0:
            paired = []
            for i in range(0, len(parts), 2):
                paired.append(f"{parts[i]}, {parts[i + 1]}")
            return paired

    return [text]


def split_display_names(value):
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []

    pattern = (
        r"[A-Za-z\u00c4\u00d6\u00dc\u00e4\u00f6\u00fc\u00df-]+"
        r",\s*[A-Za-z\u00c4\u00d6\u00dc\u00e4\u00f6\u00fc\u00df-]+"
    )
    matches = [match.strip() for match in re.findall(pattern, text) if match.strip()]
    if len(matches) >= 2:
        return matches

    return [name.strip() for name in split_names(text) if name and name.strip()]


def split_display_rooms(value):
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = re.split(r"[;,\n\r/]+|\s+", text)
    return [part.strip() for part in parts if part and part.strip()]


def name_candidates(value):
    names = split_names(value)
    return [normalize_name(name) for name in names if name]


def matches_name(cell_value, selected_name):
    if not selected_name:
        return False
    target = normalize_name(selected_name)
    return target in name_candidates(cell_value)


def extract_aufsichten(rows):
    seen = {}
    for row in rows:
        for name in split_names(row.get("Aufsicht")):
            norm = normalize_name(name)
            if norm and norm not in seen:
                seen[norm] = name
    return [seen[key] for key in sorted(seen.keys())]


def filter_rows_by_aufsicht(rows, selected_name):
    return [row for row in rows if matches_name(row.get("Aufsicht"), selected_name)]


def display_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value)


def preview_rows(rows):
    formatted = []
    for row in rows:
        formatted_row = {}
        for key, val in row.items():
            if key == "Datum":
                try:
                    formatted_row[key] = parse_date_value(val).isoformat()
                    continue
                except ValueError:
                    pass
            if key == "Startzeit":
                try:
                    formatted_row[key] = parse_time_value(val).strftime("%H:%M")
                    continue
                except ValueError:
                    pass
            if key == "Dauer":
                try:
                    formatted_row[key] = str(parse_duration_minutes(val))
                    continue
                except ValueError:
                    pass
            formatted_row[key] = display_value(val)
        formatted.append(formatted_row)
    return formatted


def parse_date_value(value):
    if value is None:
        raise ValueError("Datum fehlt")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except Exception as exc:
            raise ValueError("Datum ungueltig") from exc

    text = str(value).strip()
    if not text:
        raise ValueError("Datum fehlt")

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    try:
        return date.fromisoformat(text)
    except Exception as exc:
        raise ValueError("Datum ungueltig") from exc


def parse_time_value(value):
    if value is None:
        raise ValueError("Startzeit fehlt")
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, (int, float)):
        seconds = round(float(value) * 86400)
        return (datetime(1900, 1, 1) + timedelta(seconds=seconds)).time()

    text = str(value).strip()
    if not text:
        raise ValueError("Startzeit fehlt")

    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue

    try:
        return time.fromisoformat(text)
    except ValueError:
        try:
            return datetime.fromisoformat(text).time()
        except Exception as exc:
            raise ValueError("Startzeit ungueltig") from exc


def parse_duration_minutes(value):
    if value is None:
        raise ValueError("Dauer fehlt")
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.hour * 60 + value.minute + round(value.second / 60)
    if isinstance(value, timedelta):
        return int(round(value.total_seconds() / 60))
    if isinstance(value, (int, float)):
        number = float(value)
        if 0 < number < 1:
            return int(round(number * 1440))
        return int(round(number))

    text = str(value).strip()
    if not text:
        raise ValueError("Dauer fehlt")

    if ":" in text:
        parts = text.split(":")
        if len(parts) >= 2:
            hours = int(parts[0])
            minutes = int(parts[1])
            return hours * 60 + minutes

    try:
        return int(round(float(text)))
    except Exception as exc:
        raise ValueError("Dauer ungueltig") from exc


def prepare_event_data(row):
    pruefungsname = display_value(row.get("Pr\u00fcfungsname")).strip()
    pruefer = display_value(row.get("Pr\u00fcfer")).strip()
    aufsicht = display_value(row.get("Aufsicht")).strip()
    abloesung = display_value(row.get("Abl\u00f6sung")).strip()
    raum = display_value(row.get("Raum")).strip()

    datum = parse_date_value(row.get("Datum"))
    startzeit = parse_time_value(row.get("Startzeit"))
    dauer_minuten = parse_duration_minutes(row.get("Dauer"))

    return {
        "pruefungsname": pruefungsname,
        "pruefer": pruefer,
        "aufsicht": aufsicht,
        "abloesung": abloesung,
        "raum": raum,
        "datum": datum,
        "startzeit": startzeit,
        "dauer_minuten": dauer_minuten,
    }


def row_fingerprint(event_data):
    parts = [
        event_data.get("pruefungsname", ""),
        event_data.get("datum").isoformat(),
        event_data.get("startzeit").strftime("%H:%M"),
        str(event_data.get("dauer_minuten")),
        event_data.get("raum", ""),
        event_data.get("aufsicht", ""),
        event_data.get("abloesung", ""),
    ]
    joined = "|".join(part.strip() for part in parts)
    return hashlib.sha256(joined.encode("utf-8")).hexdigest()


def _detect_header_row(rows, scan_limit=30):
    best_index = 0
    best_headers = []
    best_score = -1

    for index, row in enumerate(rows[:scan_limit]):
        headers = [str(value).strip() if value is not None else "" for value in row]
        score = sum(1 for field in EXPECTED_COLUMNS if _resolve_indices(headers, field))
        if score > best_score:
            best_index = index
            best_headers = headers
            best_score = score

    return best_index, best_headers, best_score


def _normalize_header(value):
    text = " ".join(str(value or "").strip().split())
    text = text.replace(" / ", "/").replace(" /", "/").replace("/ ", "/")
    text = (
        text.replace("\u00e4", "ae")
        .replace("\u00f6", "oe")
        .replace("\u00fc", "ue")
        .replace("\u00df", "ss")
    )
    return text.casefold()


def _find_indices_by_names(headers, names):
    targets = {_normalize_header(name) for name in names}
    indices = []
    for idx, header in enumerate(headers):
        if _normalize_header(header) in targets:
            indices.append(idx)
    return indices


def _resolve_indices(headers, field):
    indices = _find_indices_by_names(headers, COLUMN_ALIASES.get(field, []))
    if indices:
        return indices

    if field == "Pr\u00fcfer":
        return [
            idx
            for idx, header in enumerate(headers)
            if _normalize_header(header).startswith("pruefer")
        ]

    if field == "Raum":
        prefixes = ("raum", "raeume")
        return [
            idx
            for idx, header in enumerate(headers)
            if _normalize_header(header).startswith(prefixes)
        ]

    return []


def _first_value_by_alias(headers, row, aliases):
    for alias in aliases:
        for idx in _find_indices_by_names(headers, [alias]):
            value = row[idx] if idx < len(row) else None
            if value is None:
                continue
            if str(value).strip():
                return value
    return None


def _collect_values(headers, row, indices):
    values = []
    seen = set()
    for idx in indices:
        value = row[idx] if idx < len(row) else None
        if value is None:
            continue
        text = str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        values.append(text)
    return ", ".join(values) if values else None


def _build_record(headers, row):
    record = {}
    for field in EXPECTED_COLUMNS:
        aliases = COLUMN_ALIASES.get(field, [field])
        if field == "Pr\u00fcfer":
            indices = _resolve_indices(headers, field)
            record[field] = _collect_values(headers, row, indices)
            continue
        if field == "Raum":
            indices = _resolve_indices(headers, field)
            record[field] = _collect_values(headers, row, indices)
            continue

        value = _first_value_by_alias(headers, row, aliases)
        record[field] = value

    if all(not str(value).strip() for value in record.values() if value is not None):
        return None

    return record


def _load_workbook(file_path):
    try:
        return load_workbook(file_path, data_only=True)
    except Exception:
        return load_workbook(file_path, data_only=True, read_only=True, keep_links=False)


def _read_rows_openpyxl(file_path):
    workbook = _load_workbook(file_path)
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))


def _read_rows_fallback(file_path):
    with zipfile.ZipFile(file_path) as archive:
        sheet_path = _find_sheet_path(archive)
        shared_strings = _read_shared_strings(archive)
        xml_bytes = archive.read(sheet_path)

    return _parse_sheet_rows(xml_bytes, shared_strings)


def _find_sheet_path(archive):
    names = set(archive.namelist())
    if "xl/worksheets/sheet1.xml" in names:
        return "xl/worksheets/sheet1.xml"

    sheet_names = sorted(
        name
        for name in names
        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
    )
    if not sheet_names:
        raise ValueError("Keine Arbeitsblaetter gefunden")
    return sheet_names[0]


def _read_shared_strings(archive):
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []

    xml_bytes = archive.read("xl/sharedStrings.xml")
    root = ET.fromstring(xml_bytes)
    strings = []
    for si in root.findall(".//{*}si"):
        parts = [
            text_node.text or ""
            for text_node in si.findall(".//{*}t")
            if text_node.text
        ]
        strings.append("".join(parts))
    return strings


def _parse_sheet_rows(xml_bytes, shared_strings):
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(".//{*}sheetData")
    if sheet_data is None:
        return []

    rows = []
    for row in sheet_data.findall("{*}row"):
        row_values = {}
        for cell in row.findall("{*}c"):
            cell_ref = cell.attrib.get("r", "")
            col_index = _col_to_index(cell_ref)
            if col_index is None:
                continue

            value = _read_cell_value(cell, shared_strings)
            row_values[col_index] = value

        if not row_values:
            continue

        max_idx = max(row_values)
        row_list = [None] * (max_idx + 1)
        for idx, value in row_values.items():
            row_list[idx] = value
        rows.append(row_list)

    return rows


def _read_cell_value(cell, shared_strings):
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        value_node = cell.find("{*}v")
        if value_node is None or value_node.text is None:
            return None
        index = int(value_node.text)
        if index < len(shared_strings):
            return shared_strings[index]
        return None

    if cell_type == "inlineStr":
        text_node = cell.find(".//{*}t")
        return text_node.text if text_node is not None else ""

    value_node = cell.find("{*}v")
    if value_node is None or value_node.text is None:
        return None

    text = value_node.text
    if cell_type == "b":
        return text == "1"

    return _coerce_number(text)


def _coerce_number(text):
    try:
        return int(text)
    except ValueError:
        try:
            return float(text)
        except ValueError:
            return text


def _col_to_index(cell_ref):
    match = re.match(r"([A-Za-z]+)", cell_ref)
    if not match:
        return None

    letters = match.group(1).upper()
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1

